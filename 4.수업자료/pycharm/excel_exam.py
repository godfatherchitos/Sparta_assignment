from openpyxl import load_workbook

work_book = load_workbook('prac01.xlsx')
work_sheet = work_book['prac']

print(work_sheet.cell(row = 1, column = 1).value)

work_sheet.cell(row=2, column = 2, value = 'hyowon')

work_book.save('prac01.xlsx')

